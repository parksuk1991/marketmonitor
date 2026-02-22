"""
섹터 ETF 엑셀 리포트 생성기
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import pandas as pd
import numpy as np

class SectorETFExcelGenerator:
    """섹터 ETF 엑셀 리포트 생성"""
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_sector_report(self, analyzed_news: List[Dict], 
                               sector_holdings: Dict, date_str: str) -> str:
        """섹터 리포트 생성"""
        
        print("\n엑셀 생성 시작...")
        print(f"  분석된 뉴스: {len(analyzed_news)}개")
        
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        # 시트 1: Daily News Monitor
        ws_news = wb.create_sheet("Daily News Monitor", 0)
        self._create_news_sheet(ws_news, analyzed_news, sector_holdings)
        
        # 시트 2: Sentiment Trend (선택사항)
        ws_trend = wb.create_sheet("Sentiment Trend", 1)
        self._create_trend_sheet(ws_trend, analyzed_news)
        
        # 파일 저장
        filename = f"Market_Monitor_{date_str}.xlsx"
        filepath = self.output_dir / filename
        
        wb.save(filepath)
        
        print(f"✅ 엑셀 저장: {filepath}")
        
        return str(filepath)
    
    def _create_news_sheet(self, ws, analyzed_news: List[Dict], 
                          sector_holdings: Dict):
        """Daily News Monitor 시트 생성"""
        
        # 헤더
        headers = [
            'ETF', 'Sector', 'Ticker', 'Company', 'Weight (%)',
            'Category', 'Title', 'URL', 'Pub Date', 'Highlights', 'Sentiment'
        ]
        
        ws.append(headers)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 섹터별로 정리
        print("섹터별 뉴스 분류...")
        news_by_sector = {}
        for news in analyzed_news:
            sector = news.get('sector', 'Unknown')
            if sector not in news_by_sector:
                news_by_sector[sector] = []
            news_by_sector[sector].append(news)
        
        # 섹터별 점수 계산
        print("섹터별 점수 계산...")
        sector_scores = {}
        for sector, news_list in news_by_sector.items():
            sentiments = [n.get('sentiment_score', 0.0) for n in news_list]
            
            if sentiments:
                # Simple Average
                simple_avg = np.mean(sentiments)
                
                # Weighted Average (상위 종목에 더 높은 가중치)
                weights = [n.get('weight', 1.0) for n in news_list]
                if sum(weights) > 0:
                    weighted_avg = np.average(sentiments, weights=weights)
                else:
                    weighted_avg = simple_avg
                
                sector_scores[sector] = {
                    'simple': simple_avg,
                    'weighted': weighted_avg,
                    'count': len(sentiments)
                }
                
                print(f"  {sector}:")
                print(f"    Simple: {simple_avg:.4f}")
                print(f"    Weighted: {weighted_avg:.4f}")
        
        # 데이터 입력 (섹터별 그룹화)
        row_num = 2
        
        for sector in sorted(news_by_sector.keys()):
            news_list = news_by_sector[sector]
            
            # 섹터 정보가 있으면
            sector_info = sector_holdings.get(sector, {})
            etf = sector_info.get('etf', '')
            
            # 섹터 헤더 행
            if sector in sector_scores:
                scores = sector_scores[sector]
                
                ws.cell(row_num, 1, etf)  # ETF
                ws.cell(row_num, 2, sector)  # Sector
                ws.cell(row_num, 3, f"Simple: {scores['simple']:.4f}")  # Ticker에 Simple 점수
                ws.cell(row_num, 4, f"Weighted: {scores['weighted']:.4f}")  # Company에 Weighted 점수
                
                # 헤더 행 스타일
                for col in range(1, 5):
                    cell = ws.cell(row_num, col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                
                row_num += 1
            
            # 뉴스 데이터
            for news in news_list:
                row_data = [
                    etf,
                    sector,
                    news.get('ticker', ''),
                    news.get('company_name', ''),
                    news.get('weight', 0.0),
                    news.get('category', 'General'),
                    news.get('title', ''),
                    news.get('url', ''),
                    news.get('published_at', '')[:10],
                    news.get('summary', '')[:100] + '...' if news.get('summary') else '',
                    news.get('sentiment_score', 0.0)
                ]
                
                ws.append(row_data)
                
                # Sentiment 색상
                sentiment_cell = ws.cell(row_num, 11)
                self._apply_sentiment_color(sentiment_cell, news.get('sentiment_score', 0.0))
                
                # URL 하이퍼링크
                url_cell = ws.cell(row_num, 8)
                if news.get('url'):
                    url_cell.hyperlink = news['url']
                    url_cell.font = Font(color='0563C1', underline='single')
                
                row_num += 1
        
        # 열 너비 조정
        column_widths = [10, 25, 12, 25, 10, 12, 60, 15, 12, 50, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 필터 추가
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        
        print(f"✅ 메인 시트 완료: {row_num-1}행")
    
    def _create_trend_sheet(self, ws, analyzed_news: List[Dict]):
        """Sentiment Trend 시트 생성"""
        
        # 티커별로 그룹화
        ticker_sentiments = {}
        
        for news in analyzed_news:
            ticker = news.get('ticker', '')
            if not ticker:
                continue
            
            if ticker not in ticker_sentiments:
                ticker_sentiments[ticker] = {
                    'company': news.get('company_name', ''),
                    'sector': news.get('sector', ''),
                    'dates': {}
                }
            
            date = news.get('published_at', '')[:10]
            sentiment = news.get('sentiment_score', 0.0)
            
            if date not in ticker_sentiments[ticker]['dates']:
                ticker_sentiments[ticker]['dates'][date] = []
            
            ticker_sentiments[ticker]['dates'][date].append(sentiment)
        
        # 헤더
        headers = ['Ticker', 'Company', 'Sector', 'Date -2', 'Date -1', 'Today', 'Trend', 'Change']
        ws.append(headers)
        
        # 헤더 스타일
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 데이터
        row_num = 2
        dates = sorted(set(date for data in ticker_sentiments.values() 
                          for date in data['dates'].keys()), reverse=True)[:3]
        
        for ticker, data in sorted(ticker_sentiments.items()):
            if len(dates) < 2:
                continue
            
            # 각 날짜별 평균
            date_avgs = []
            for date in reversed(dates):  # 오래된 날짜부터
                if date in data['dates']:
                    avg = np.mean(data['dates'][date])
                    date_avgs.append(avg)
                else:
                    date_avgs.append(None)
            
            # 최소 2개 날짜 필요
            if len([x for x in date_avgs if x is not None]) < 2:
                continue
            
            # 트렌드 계산
            valid_values = [x for x in date_avgs if x is not None]
            if len(valid_values) >= 2:
                change = valid_values[-1] - valid_values[0]
                trend = "📈" if change > 0.1 else "📉" if change < -0.1 else "➡️"
            else:
                change = 0.0
                trend = "➡️"
            
            # 행 데이터
            row_data = [
                ticker,
                data['company'],
                data['sector']
            ]
            
            # 날짜별 값 (최대 3개)
            for i in range(3):
                if i < len(date_avgs):
                    row_data.append(date_avgs[i] if date_avgs[i] is not None else '')
                else:
                    row_data.append('')
            
            row_data.extend([trend, change])
            
            ws.append(row_data)
            row_num += 1
        
        # 열 너비
        column_widths = [12, 25, 20, 12, 12, 12, 8, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        print(f"✅ 트렌드 시트 완료: {row_num-1}행")
    
    def _apply_sentiment_color(self, cell, sentiment: float):
        """Sentiment에 따른 색상 적용"""
        if sentiment > 0.2:
            # 긍정: 초록
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            cell.font = Font(color='006100')
        elif sentiment < -0.2:
            # 부정: 빨강
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(color='9C0006')
        else:
            # 중립: 노랑
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            cell.font = Font(color='9C6500')
        
        cell.number_format = '0.0000'
        cell.alignment = Alignment(horizontal='center')
